from tkinter import *
from tkinter import filedialog
import os
from tkinter import messagebox
from PIL import Image, ImageTk
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib


# Function for saving user input to MS Excel and informs user if the data was saved, including user image and clears all field when the command is performed.
def register():
    getName = entry_name.get()
    getAge = entry_age.get()
    getAddress = entry_address.get()
    getGender = entry_gender.get()
    getBirthday = entry_birth.get()
    getEmail = entry_email.get()
    getContact = entry_contact.get()
    getYear = entry_year.get()
    getUnits = entry_units.get()
    getCourse = entry_course.get()

    if getName == "Enter your name" or getAge == "Enter your age" or getAddress == "Enter your current address" or getGender == "" \
            or getBirthday == "Date of Birth (mm/dd/yy)" or getEmail == "E-mail Address" or getContact == "Contact Number" \
            or getYear == "Select Year Level" or getUnits == "Select No. of Units" or getCourse == "Select Program" or image1 == "xx.png":
        messagebox.showerror("Error", "All information must be filled up")
    else:
        file = openpyxl.load_workbook("Student Information.xlsx")
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=getName)
        sheet.cell(column=2, row=sheet.max_row, value=getAge)
        sheet.cell(column=3, row=sheet.max_row, value=getAddress)
        sheet.cell(column=4, row=sheet.max_row, value=getGender)
        sheet.cell(column=5, row=sheet.max_row, value=getBirthday)
        sheet.cell(column=6, row=sheet.max_row, value=getContact)
        sheet.cell(column=7, row=sheet.max_row, value=getEmail)
        sheet.cell(column=8, row=sheet.max_row, value=getYear)
        sheet.cell(column=9, row=sheet.max_row, value=getCourse)
        sheet.cell(column=10, row=sheet.max_row, value=getUnits)
        file.save(r"Student Information.xlsx")

        imageIn.save("Student Images/" + getName + ".jpg")
        messagebox.showinfo("Notice", "Register Success! Data has been saved.")
        clear()


# Function for clearing all fields, including user image, radiobutton, and combobox.
def clear():
    global imageIn
    entry_name.set("Enter your name")
    entry_age.set("Enter your age")
    entry_address.set("Enter your current address")
    entry_birth.set("Date of Birth (mm/dd/yy)")
    entry_contact.set("Contact Number")
    entry_year.set("Select Year Level")
    entry_units.set("Select No. of Units")
    entry_course.set("Select Program")
    entry_email.set("E-mail Address")
    male.deselect()
    female.deselect()

    registerButton.config(state="normal")
    image1 = ImageTk.PhotoImage(file="pyimage\\FrameImage.png")
    label.image = image1


# Function to exit the program.
def exit():
    window.destroy()


# Function for uploading user image on frame.
def upload_image():
    global filename
    global imageIn
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select Image", filetypes=(("JPG File", "*jpg"), ("PNG File", "*png"),
                                                                           ["All Files", "*.txt"]))
    imageIn = (Image.open(filename))
    resize_image = imageIn.resize((89, 89))
    image = ImageTk.PhotoImage(resize_image)
    label.config(image=image)
    label.image = image


# The Window Canvas (includes the name, width, height, and background)
bg_color = "#fff"
window = Tk()
window.title("Enrollment/Registration Form")
window.geometry("1050x1000")
window.config(bg=bg_color)

# Personal Information Input
entry_name = StringVar()
entry_age = StringVar()
entry_address = StringVar()
entry_birth = StringVar()
entry_gender = StringVar()
entry_contact = StringVar()
entry_year = StringVar()
entry_units = StringVar()
entry_course = StringVar()
entry_email = StringVar()

# Field Text
Label(window, text="", width=70, height=500, anchor="e", bg="#385723").pack(side=LEFT)
Label(window, text="", width=150, height=8, anchor="e", bg="#f5f5dc").pack(side=TOP)
Label(text="Registration Form", bg="#385723", fg="white", font="times 26").place(x=112, y=50)


def enter(e):
    user.delete(0, "end")


def leave(e):
    if user.get() == "":
        user.insert(0, "Enter your name")


user = Entry(window, width=25, textvariable=entry_name, fg="#383838", border=0, bg="white", font=("verdana", 13))
user.place(x=550, y=175)
Frame(window, width=270, height=2, bg="black").place(x=545, y=200)
user.insert(0, "Enter your name")

user.bind("<FocusIn>", enter)
user.bind("<FocusOut>", leave)


def enter(e):
    age.delete(0, "end")


def leave(e):
    if age.get() == "":
        age.insert(0, "Enter your age")


age = Entry(window, width=25, textvariable=entry_age, border=0, bg="white", fg="#383838", font=("verdana", 13))
age.place(x=550, y=240)
Frame(window, width=270, height=2, bg="black").place(x=545, y=265)
age.insert(0, "Enter your age")
age.bind("<FocusIn>", enter)
age.bind("<FocusOut>", leave)


def enter(e):
    address.delete(0, "end")


def leave(e):
    if address.get() == "":
        address.insert(0, "Enter your current address")


address = Entry(window, width=25, textvariable=entry_address, fg="#383838", border=0, bg="white", font=("verdana", 13))
address.place(x=550, y=305)
Frame(window, width=270, height=2, bg="black").place(x=545, y=330)
address.insert(0, "Enter your current address")
address.bind("<FocusIn>", enter)
address.bind("<FocusOut>", leave)


def selection():
    value = entry_gender.get()
    if value != "Female":
        print(value)
    else:
        print(value)


Label(text="Gender", bg="white", fg="#383838", font="verdana 13").place(x=545, y=370)
male = Radiobutton(window, text="Male", font="verdana 10", command=selection, variable=entry_gender, value="Male",
                   bg="white", fg="black"
                   )
male.place(x=650, y=370)
female = Radiobutton(window, text="Female", font="verdana 10", command=selection, variable=entry_gender, value="Female",
                     bg="white", fg="black"
                     )
female.place(x=730, y=370)


def enter(e):
    birthday.delete(0, "end")


def leave(e):
    if birthday.get() == "":
        birthday.insert(0, "Date of Birth (mm/dd/yy)")


birthday = Entry(window, width=25, textvariable=entry_birth, fg="#383838", border=0, bg="white", font=("verdana", 13))
birthday.place(x=550, y=435)
Frame(window, width=270, height=2, bg="black").place(x=545, y=460)
birthday.insert(0, "Date of Birth (mm/dd/yy)")
birthday.bind("<FocusIn>", enter)
birthday.bind("<FocusOut>", leave)


def enter(e):
    email.delete(0, "end")


def leave(e):
    if email.get() == "":
        email.insert(0, "E-mail Address")


email = Entry(window, width=25, textvariable=entry_email, fg="#383838", border=0, bg="white", font=("verdana", 13))
email.place(x=550, y=500)
Frame(window, width=270, height=2, bg="black").place(x=545, y=525)
email.insert(0, "E-mail Address")
email.bind("<FocusIn>", enter)
email.bind("<FocusOut>", leave)


def enter(e):
    contact.delete(0, "end")


def leave(e):
    if contact.get() == "":
        contact.insert(0, "Contact Number")


contact = Entry(window, width=25, textvariable=entry_contact, fg="#383838", border=0, bg="white", font=("verdana", 13))
contact.place(x=550, y=565)
Frame(window, width=270, height=2, bg="black").place(x=545, y=590)
contact.insert(0, "Contact Number")
contact.bind("<FocusIn>", enter)
contact.bind("<FocusOut>", leave)

Label(text="Year Level:", bg="white", fg="#383838", font="verdana 13").place(x=900, y=171)
year = Combobox(window, values=["1st Year", "2nd Year", "3rd Year", "4th Year", "5th Year"], textvariable=entry_year,
                font="verdana 13", width=17, state="r")
year.place(x=1025, y=171)
year.set("Select Year Level")

Label(text="No. of Units:", bg="white", fg="#383838", font="verdana 13").place(x=900, y=236)
units = Combobox(window, values=["12", "13", "14", "15", "16", "17", "18", "19", "20", "21", "22", "23", "24", ],
                 textvariable=entry_units, font="verdana 13", width=17, state="r")
units.place(x=1025, y=236)
units.set("Select No. of Units")

Label(text="Program:", bg="white", fg="#383838", font="verdana 13").place(x=900, y=301)
course = Combobox(window, values=["Bachelor of Secondary Education", "BS Business Management",
                                  "BS Hotel and Restaurant Management", "BS Information Technology",
                                  "BS Office Administration", "BS Psychology"], textvariable=entry_course,
                  font="verdana 13", width=17, state="r")
course.place(x=1025, y=301)
course.set("Select Program")

photo = Frame(window, bd=3, bg="white", width=100, height=100, relief=GROOVE)
photo.place(x=545, y=10)
imageIn = ImageTk.PhotoImage(file="pyimage\\FrameImage.png")
label = Label(photo, bg="white", image=imageIn)
label.place(x=0, y=0)

# Image Inputs
image1 = ImageTk.PhotoImage(file="pyimage\\cvsu4.png")
Label(window, image=image1, border=0, bg="white").place(x=10, y=115)

Button(window, text="Upload", width=11, height=1,
       command=upload_image, font="verdana 9 bold", bg="#385723", fg="white").place(x=545, y=100)

registerButton = Button(window, text="Register", font="verdana 14 bold", width=15, height=2, bg="#385723", fg="white",
                        command=register)
registerButton.place(x=1025, y=375)
Button(window, text="Exit", bg="#f5f5dc", bd=0, fg="#385723", command=exit, font="verdana 13 bold underline").place(
    x=1200, y=90)
Button(window, text="Reset All Field", font="verdana 14 bold underline", bd=0, width=15, height=1, bg="white",
       fg="#385723", command=clear).place(x=1025,
                                          y=450)

# Ecxel Database
file = pathlib.Path("Student Information.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Student Name"
    sheet["B1"] = "Age"
    sheet["C1"] = "Complete Address"
    sheet["D1"] = "Gender"
    sheet["E1"] = "Date of Birth"
    sheet["F1"] = "Contact Number"
    sheet["G1"] = "E-mail Address"
    sheet["H1"] = "Year Level"
    sheet["I1"] = "Program"
    sheet["J1"] = "Number of Units"
    file.save("Student Information.xlsx")

window.mainloop()
